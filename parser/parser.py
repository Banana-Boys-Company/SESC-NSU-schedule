import openpyxl as xl
from openpyxl.utils.exceptions import SheetTitleException
import re
from openpyxl.cell.cell import Cell
from json import dumps, dump
import copy
from common import pattern, dow2dow, subj, validate_str, get_merged_cell_val, full_subj_name, first_table_properties, \
    ScheduleProperties, second_table_properties
from typing import IO
from os.path import exists
import io


class ScheduleParser:
    wb: xl.Workbook
    prop: ScheduleProperties

    def __init__(self, table_path: str, _prop: ScheduleProperties):
        self.prop = _prop
        self.refresh(table_path)

    def refresh(self, table_path: str):
        if exists(table_path):
            self.wb: xl.Workbook = xl.load_workbook(table_path)
        else:
            raise FileNotFoundError('File does not exist')

    def parse(self, sheet_name: str, fp: IO[str] = None):
        # Loading xlsx file
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            raise SheetTitleException('Sheet with this name does not exist')
        if not re.match(r'[+w]t?', fp.mode):
            raise io.UnsupportedOperation('File is not writable')

        # Initiating variables
        classes = dict()
        same_lesson = False
        times = []
        row_to_day_of_the_week = dict()
        previous_day = ''

        # Generating dictionary structure
        for i, column in enumerate(ws.iter_cols(max_row=2)):
            __v = str(get_merged_cell_val(ws, column[0]))
            __val = validate_str(get_merged_cell_val(ws, column[1]))
            if len(__val) < 3:
                __val = '0' * (3 - len(__val)) + __val
            if not re.match(r'\d+_\d+', __v):
                continue
            classes.setdefault(__v, dict())
            classes[__v].setdefault(__val, dict())

        # They`re not needed anymore
        del __v, i, __val, column

        # Starting parsing data
        for i, column in enumerate(ws.iter_cols(min_row=1, max_row=self.prop.last_row)):
            # First column. Data - days of week. Creating dictonary {row: day_of_week}
            if i == 0:
                for row_, cell_ in enumerate(column[2:]):
                    value = get_merged_cell_val(ws, cell_)
                    if type(cell_) == Cell and (cell_.value is not None):
                        if not [s for s in ws.merged_cells.ranges if cell_.coordinate in s]:
                            continue
                    if value is None:
                        row_to_day_of_the_week: dict
                        value = row_to_day_of_the_week[list(row_to_day_of_the_week)[-1]]
                        row_to_day_of_the_week[cell_.row] = value
                        continue
                    row_to_day_of_the_week[cell_.row] = dow2dow[value]
            # Useless columns
            elif i in self.prop.useless_columns:
                continue
            # Third column. Data - time of lessons. Not used
            elif (i == 2) and self.prop.second_time:
                for row_ in range(2, 13, 2):
                    times.append(str(column[row_].value)[:-3])
            else:
                # Creating local copy of dictionary with days of week as keys
                __dict = copy.deepcopy(pattern)
                if column[0].value is None and type(column[0]) == Cell:
                    continue

                # Getting class and group id
                class_num = validate_str(get_merged_cell_val(ws, column[0]))
                group_num = validate_str(get_merged_cell_val(ws, column[1]))

                if len(group_num) < 3:
                    group_num = '0' * (3 - len(group_num)) + group_num

                # Looking through every cell
                for row, cell in enumerate(column[2:]):
                    # Getting value of cell
                    val = validate_str(get_merged_cell_val(ws, cell))

                    # Filtrating empty rows
                    if val == 'None':
                        if cell.row not in self.prop.useless_rows:
                            val = '\0'
                        else:
                            continue
                    if cell.row in self.prop.useless_rows:
                        continue

                    # Getting width of row (bugged ones count as 2)
                    rowlen = 2 if cell.row in self.prop.bug_rows else 1

                    # Get actual day of week
                    day = row_to_day_of_the_week[cell.row]

                    _val = str(val).rstrip('.').upper()
                    if _val in subj:
                        val = full_subj_name[_val]

                    if row == 0:
                        __dict[day] += [[val, rowlen]]
                    # Check if new day has come
                    elif day != previous_day:
                        __dict[day] += [[val, rowlen]]
                    else:
                        # Checking if this row value is the same as the previous one (multiple rows for one lesson)
                        if val in __dict[day][-1][0] and _val not in subj:
                            __dict[day][-1][1] += rowlen
                        else:
                            # Checking if there`s two separates rows with different values for one lesson
                            if same_lesson:
                                __dict[day][-1][0] += f' {val}'
                                __dict[day][-1][1] += rowlen
                            else:
                                __dict[day] += [[val, rowlen]]
                    # Cheking if this row was the first part of multiple rows for one lesson
                    same_lesson = str(_val).strip('.').upper() in subj
                    previous_day = day

                # Divide every lesson length (in rows) to get actual length of lesson
                for __day in __dict:
                    _day = __dict[__day]
                    for __lesson in _day:
                        __lesson[1] //= 2

                # Writing data in main dictionary
                classes[class_num][group_num] = __dict

        # Returning json with unicode characters
        if fp:
            dump(classes, fp=fp, ensure_ascii=False)
        return dumps(classes, ensure_ascii=False)


if __name__ == '__main__':
    p = ScheduleParser('SESC_Timetable 2022_2023.xlsx', second_table_properties)
    with open('json.json', 'w', encoding='utf8') as file:
        json = p.parse('Расписание_1сем_2пол.дня', fp=file)
