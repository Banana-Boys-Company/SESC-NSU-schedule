import openpyxl as xl
from openpyxl.utils.exceptions import SheetTitleException
import re
from openpyxl.cell.cell import Cell
from json import dump
import copy
from typing import IO
from os.path import exists
import io
import progress.bar
try:
    from modules.parser.common import pattern, dow2dow, subj, validate_str, get_merged_cell_val, full_subj_name, \
        first_table_properties, ScheduleProperties, second_table_properties, merge_dicts, ProgressPlug, department_to_id
except ImportError:
    from modules.parser.common import pattern, dow2dow, subj, validate_str, get_merged_cell_val, full_subj_name, \
        first_table_properties, ScheduleProperties, second_table_properties, merge_dicts, ProgressPlug, department_to_id


class ScheduleParser:
    wb: xl.Workbook
    prop: ScheduleProperties

    def __init__(self, table_path: str):
        self.refresh(table_path)

    def __set_prop(self, _prop: ScheduleProperties):
        self.prop = _prop

    def refresh(self, table_path: str):
        if exists(table_path):
            self.wb: xl.Workbook = xl.load_workbook(table_path)
        else:
            raise FileNotFoundError('File does not exist')
        print('File loaded!')

    def parse(self, sheet_name: str, _prop: ScheduleProperties, fp: IO[str] = None, bar_is_on: bool = False):

        # region Loading xlsx file
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            raise SheetTitleException('Sheet with this name does not exist')
        if fp:
            if not re.match(r'[+w]t?', fp.mode):
                raise io.UnsupportedOperation('File is not writable')
        # endregion
        # region Initiating variables
        self.__set_prop(_prop)
        classes = dict()
        same_lesson = False
        times = []
        row_to_day_of_the_week = dict()
        previous_day = ''
        if bar_is_on:
            bar = progress.bar.ChargingBar(
                f'Parsing {sheet_name}', max=_prop.progress)
        else:
            bar = ProgressPlug
        # endregion
        # region Generating dictionary structure
        for i, column in enumerate(ws.iter_cols(max_row=2)):
            __v = str(get_merged_cell_val(ws, column[0]))
            __val = validate_str(get_merged_cell_val(ws, column[1]))
            if len(__val) < 3:
                __val = '0' * (3 - len(__val)) + __val
            if not re.match(r'\d+_\d+', __v):
                continue
            classes.setdefault(__v, dict())
            classes[__v].setdefault(__val, dict())
            bar.next()
        # endregion
        # They`re not needed anymore
        del __v, i, __val, column

        # region Parsing data
        for i, column in enumerate(ws.iter_cols(min_row=1, max_row=self.prop.last_row)):
            # region First column. Data - days of week. Creating dictonary {row: day_of_week}
            if i == 0:
                for row_, cell_ in enumerate(column[2:]):
                    bar.next()
                    value = get_merged_cell_val(ws, cell_)
                    if type(cell_) == Cell and (cell_.value is not None):
                        if not [s for s in ws.merged_cells.ranges if cell_.coordinate in s]:
                            continue
                    if value is None:
                        row_to_day_of_the_week: dict
                        value = row_to_day_of_the_week[list(
                            row_to_day_of_the_week)[-1]]
                        row_to_day_of_the_week[cell_.row] = value
                        continue
                    row_to_day_of_the_week[cell_.row] = dow2dow[value]
            # endregion
            # region Useless columns
            elif i in self.prop.useless_columns:
                continue
            # endregion
            # region Third column. Data - time of lessons. Not used
            elif (i == 2) and self.prop.second_time:
                for row_ in range(2, 13, 2):
                    times.append(str(column[row_].value)[:-3])
            # endregion
            else:
                # Creating local copy of dictionary with days of week as keys
                __dict = copy.deepcopy(pattern)
                if column[0].value is None and type(column[0]) == Cell:
                    continue

                # Getting class and group id
                class_num = validate_str(get_merged_cell_val(ws, column[0]))
                group_num = validate_str(get_merged_cell_val(ws, column[1]))

                if len(group_num) < 3:
                    group_num = '0' * (3 - len(group_num)) + group_num

                # region Looking through every cell
                for row, cell in enumerate(column[2:]):
                    # Getting value of cell
                    val = validate_str(get_merged_cell_val(ws, cell))
                    bar.next()
                    # Filtrating empty rows
                    if val == 'None':
                        if cell.row not in self.prop.useless_rows:
                            val = '\0'
                        else:
                            continue
                    if cell.row in self.prop.useless_rows:
                        continue

                    # Getting width of row (bugged ones count as 2)
                    rowlen = 2 if cell.row in self.prop.bug_rows else 1

                    # Get actual day of week
                    day = row_to_day_of_the_week[cell.row]

                    _val = str(val).rstrip('.').upper()
                    if _val in subj:
                        val = full_subj_name[_val]

                    if row == 0:
                        __dict[day] += [[val, rowlen]]
                    # Check if new day has come
                    elif day != previous_day:
                        __dict[day] += [[val, rowlen]]
                    else:
                        # Checking if this row value is the same as the previous one (multiple rows for one lesson)
                        if val in __dict[day][-1][0] and _val not in subj:
                            __dict[day][-1][1] += rowlen
                        else:
                            # Checking if there`s two separates rows with different values for one lesson
                            if same_lesson:
                                __dict[day][-1][0] += f' {val}'
                                __dict[day][-1][1] += rowlen
                            else:
                                __dict[day] += [[val, rowlen]]
                    # Cheking if this row was the first part of multiple rows for one lesson
                    same_lesson = str(_val).strip('.').upper() in subj
                    previous_day = day
                # endregion
                # region Divide every lesson length (in rows) to get actual length of lesson
                for __day in __dict:
                    _day = __dict[__day]
                    for __lesson in _day:
                        __lesson[1] //= 2
                        bar.next()
                # endregion

                # Writing data in main dictionary
                classes[class_num][group_num] = __dict
        # endregion
        bar.finish()

        # region Returning json with unicode characters
        if fp:
            dump(classes, fp=fp, ensure_ascii=False)
        return copy.deepcopy(classes)
        # endregion

    def parse_cources(self, sheet_name: str, fp: IO[str] = None):
        pat = {
            '16:00': [],
            '18:00': [],
            '20:00': [],
            '20:30': [],
        }

        patte = {
            'monday': {},
            'tuesday': {},
            'wednesday': {},
            'thursday': {},
            'friday': {},
            'saturday': {}
        }

        # region Loading xlsx file
        if sheet_name in self.wb.sheetnames:
            import openpyxl.worksheet.worksheet
            ws: xl.worksheet.worksheet.Worksheet = self.wb[sheet_name]
        else:
            raise SheetTitleException('Sheet with this name does not exist')
        if fp:
            if not re.match(r'[+w]t?', fp.mode):
                raise io.UnsupportedOperation('File is not writable')
        # endregion

        courses = dict()
        departments = []
        times = []

        for i, column in enumerate(ws.iter_cols(min_col=1, max_col=8, min_row=1, max_row=84)):
            column: list[openpyxl.cell.Cell]
            if i == 0:
                departments = list(column[1:])
                continue
            if i == 1:
                times = column[1:]
                continue
            day = dow2dow[column[0].value.capitalize()]
            for row, cell in enumerate(column[1:]):
                if row == 0:
                    continue
                department = department_to_id[get_merged_cell_val(
                    ws, departments[row])]
                if get_merged_cell_val(ws, times[row]) is None:
                    continue
                time = get_merged_cell_val(ws, times[row]).strip().rstrip()
                time = f'{time[:2]}:{time[3:]}'
                try:
                    _ = courses[department]
                except KeyError:
                    courses[department] = copy.deepcopy(patte)
                try:
                    _ = courses[department][day][time]
                except KeyError:
                    courses[department][day] = copy.deepcopy(pat)
                if cell.value is None:
                    continue
                courses[department][day][time].append(cell.value)
        if fp:
            dump(courses, fp=fp, ensure_ascii=False)
        return courses


if __name__ == '__main__':
    p = ScheduleParser('SESC_Timetable 2022_2023.xlsx')
    json = p.parse_cources('СПЕЦКУРСЫ')
    with open('json.json', 'w') as file:
        dump(json, fp=file, ensure_ascii=False)
